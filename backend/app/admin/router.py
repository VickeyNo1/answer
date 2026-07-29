"""管理员模块：学生 CRUD + Excel 批量导入 + 统计数据 + 全局设置/权益 + 反馈与检索报表（v4.0）"""
import io
import json
from collections import Counter

import bcrypt
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from openpyxl import load_workbook

from app.auth.deps import require_admin
from app.database import get_db, get_db_ctx
from app.models import (
    AppSettingsOut,
    AppSettingsUpdate,
    EntitlementsUpdate,
    FeedbackItem,
    FeedbackListOut,
    HotKpItem,
    KbStatsByDay,
    KbStatsOut,
    StudentCreate,
    StudentUpdate,
    UserInfo,
    StatsResponse,
)
from app.admin.stats import get_stats
from app import settings_store

router = APIRouter(prefix="/api/admin", tags=["管理员"])


def _hash_password(password: str) -> str:
    """bcrypt 加密密码"""
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def _row_to_userinfo(row) -> dict:
    """将查询结果行转换为 UserInfo 兼容的 dict（含权益覆盖列，NULL=跟随全局）"""
    memory_enabled = row["memory_enabled"]
    return {
        "id": row["id"],
        "student_id": row["student_id"],
        "name": row["name"],
        "role": row["role"],
        "created_at": str(row["created_at"]),
        "daily_question_limit": row["daily_question_limit"],
        "memory_enabled": bool(memory_enabled) if memory_enabled is not None else None,
    }


# ========== 学生 CRUD ==========


@router.post("/students", response_model=UserInfo, status_code=status.HTTP_201_CREATED)
async def create_student(
    req: StudentCreate,
    admin: dict = Depends(require_admin),
):
    """创建学生账号"""
    with get_db_ctx() as db:
        # 检查学号是否已存在
        cursor = db.execute(
            "SELECT id FROM users WHERE student_id = %s", (req.student_id,)
        )
        if cursor.fetchone() is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"学号 {req.student_id} 已存在",
            )

        password_hash = _hash_password(req.password)
        cursor = db.execute(
            """
            INSERT INTO users (student_id, password_hash, name, role)
            VALUES (%s, %s, %s, 'student')
            """,
            (req.student_id, password_hash, req.name),
        )
        db.commit()

        # 查询新建的用户
        cursor = db.execute("SELECT * FROM users WHERE id = %s", (cursor.lastrowid,))
        return _row_to_userinfo(cursor.fetchone())


@router.get("/students")
async def list_students(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    keyword: str | None = Query(None, description="搜索学号或姓名"),
    admin: dict = Depends(require_admin),
):
    """获取学生列表（分页）"""
    with get_db_ctx() as db:
        # 构建查询条件
        where = "WHERE role = 'student'"
        params = []
        if keyword:
            where += " AND (student_id LIKE %s OR name LIKE %s)"
            kw = f"%{keyword}%"
            params.extend([kw, kw])

        # 查询总数
        cursor = db.execute(f"SELECT COUNT(*) as cnt FROM users {where}", params)
        total = cursor.fetchone()["cnt"]

        # 分页查询
        offset = (page - 1) * size
        params.extend([size, offset])
        cursor = db.execute(
            f"""
            SELECT id, student_id, name, role, created_at,
                   daily_question_limit, memory_enabled
            FROM users {where}
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
            """,
            params,
        )
        rows = cursor.fetchall()
        items = [_row_to_userinfo(row) for row in rows]

    return {
        "items": items,
        "total": total,
        "page": page,
        "size": size,
    }


@router.put("/students/{student_id}", response_model=UserInfo)
async def update_student(
    student_id: int,
    req: StudentUpdate,
    admin: dict = Depends(require_admin),
):
    """修改学生信息（name 和 password 均为可选）"""
    with get_db_ctx() as db:
        cursor = db.execute(
            "SELECT * FROM users WHERE id = %s AND role = 'student'",
            (student_id,),
        )
        row = cursor.fetchone()
        if row is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="学生不存在",
            )

        updates = []
        params = []
        if req.name is not None:
            updates.append("name = %s")
            params.append(req.name)
        if req.password is not None:
            updates.append("password_hash = %s")
            params.append(_hash_password(req.password))

        if updates:
            params.append(student_id)
            db.execute(
                f"UPDATE users SET {', '.join(updates)} WHERE id = %s",
                params,
            )
            db.commit()

        # 查询更新后的用户
        cursor = db.execute("SELECT * FROM users WHERE id = %s", (student_id,))
        return _row_to_userinfo(cursor.fetchone())


@router.delete("/students/{student_id}")
async def delete_student(
    student_id: int,
    admin: dict = Depends(require_admin),
):
    """删除学生（级联删除其对话和消息）"""
    with get_db_ctx() as db:
        cursor = db.execute(
            "SELECT * FROM users WHERE id = %s",
            (student_id,),
        )
        row = cursor.fetchone()
        if row is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="学生不存在",
            )

        if row["role"] == "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="不能删除管理员账号",
            )

        # 级联删除：先删消息，再删对话，最后删用户
        conv_ids_cursor = db.execute(
            "SELECT id FROM conversations WHERE user_id = %s",
            (student_id,),
        )
        conv_ids = [r["id"] for r in conv_ids_cursor.fetchall()]

        if conv_ids:
            placeholders = ",".join(["%s"] * len(conv_ids))
            db.execute(
                f"DELETE FROM messages WHERE conversation_id IN ({placeholders})",
                conv_ids,
            )
            db.execute(
                f"DELETE FROM conversations WHERE id IN ({placeholders})",
                conv_ids,
            )

        db.execute("DELETE FROM users WHERE id = %s", (student_id,))
        db.commit()

    return {"message": "ok"}


# ========== Excel 批量导入 ==========


@router.post("/students/batch")
async def batch_import_students(
    file: UploadFile = File(...),
    admin: dict = Depends(require_admin),
):
    """批量导入学生（Excel 格式：学号 | 姓名 | 密码）"""
    # 检查文件格式
    filename = file.filename or ""
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="仅支持 .xlsx 和 .xls 格式",
        )

    # 读取文件内容
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    success = 0
    failed = 0
    errors = []

    with get_db_ctx() as db:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            student_id = str(row[0]).strip() if row[0] else ""
            name = str(row[1]).strip() if row[1] else ""
            password = str(row[2]).strip() if len(row) > 2 and row[2] else ""

            # 校验
            if not student_id:
                failed += 1
                errors.append({"row": row_idx, "student_id": "", "reason": "学号不能为空"})
                continue
            if not name:
                failed += 1
                errors.append({"row": row_idx, "student_id": student_id, "reason": "姓名不能为空"})
                continue
            if not password:
                failed += 1
                errors.append({"row": row_idx, "student_id": student_id, "reason": "密码不能为空"})
                continue

            # 检查学号是否已存在
            cursor = db.execute(
                "SELECT id FROM users WHERE student_id = %s", (student_id,)
            )
            if cursor.fetchone() is not None:
                failed += 1
                errors.append({"row": row_idx, "student_id": student_id, "reason": "学号已存在"})
                continue

            # 创建学生
            try:
                password_hash = _hash_password(password)
                db.execute(
                    """
                    INSERT INTO users (student_id, password_hash, name, role)
                    VALUES (%s, %s, %s, 'student')
                    """,
                    (student_id, password_hash, name),
                )
                success += 1
            except Exception as e:
                failed += 1
                errors.append({"row": row_idx, "student_id": student_id, "reason": str(e)})

        db.commit()

    wb.close()

    return {
        "success": success,
        "failed": failed,
        "errors": errors,
    }


# ========== 统计数据 ==========


@router.get("/stats", response_model=StatsResponse)
async def get_admin_stats(
    admin: dict = Depends(require_admin),
):
    """获取系统统计数据"""
    return get_stats()


# ========== 全局设置（v4.0 M1） ==========


@router.get("/settings", response_model=AppSettingsOut)
async def get_app_settings(
    admin: dict = Depends(require_admin),
):
    """获取全局设置（内存缓存，值已按键转型）"""
    return settings_store.get_all()


@router.put("/settings")
async def update_app_settings(
    req: AppSettingsUpdate,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    """部分更新全局设置（只传要改的键），更新后刷新内存缓存"""
    updates = {
        key: int(value)
        for key, value in req.model_dump(exclude_unset=True).items()
        if value is not None
    }
    if updates:
        settings_store.update_settings(db, updates)
    return {"message": "ok"}


# ========== 单个学生权益（v4.0 M1） ==========


@router.put("/students/{student_id}/entitlements")
async def update_student_entitlements(
    student_id: int,
    req: EntitlementsUpdate,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    """设置单个学生权益覆盖值（null=恢复跟随全局默认）"""
    cursor = db.execute(
        "SELECT id FROM users WHERE id = %s AND role = 'student'", (student_id,)
    )
    if cursor.fetchone() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="学生不存在",
        )

    provided = req.model_dump(exclude_unset=True)
    updates = []
    params = []
    if "daily_question_limit" in provided:
        updates.append("daily_question_limit = %s")
        params.append(provided["daily_question_limit"])
    if "memory_enabled" in provided:
        updates.append("memory_enabled = %s")
        value = provided["memory_enabled"]
        params.append(int(value) if value is not None else None)

    if updates:
        params.append(student_id)
        db.execute(
            f"UPDATE users SET {', '.join(updates)} WHERE id = %s", params
        )
        db.commit()
    return {"message": "ok"}


# ========== 反馈明细（v4.0 M1） ==========


@router.get("/feedbacks", response_model=FeedbackListOut)
async def list_feedbacks(
    rating: str | None = Query(None, description="up / down 筛选"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    """反馈明细列表（按时间降序分页；「上一条学生提问」应用层关联）"""
    where = ""
    params: list = []
    if rating:
        where = "WHERE f.rating = %s"
        params.append(rating)

    cursor = db.execute(
        f"SELECT COUNT(*) AS cnt FROM feedbacks f {where}", params
    )
    total = int(cursor.fetchone()["cnt"])

    offset = (page - 1) * page_size
    cursor = db.execute(
        f"""SELECT f.id, f.rating, f.reason, f.created_at, f.message_id,
                   u.student_id, u.name AS student_name,
                   m.content AS answer, m.knowledge_point_ids, m.conversation_id
            FROM feedbacks f
            JOIN users u ON f.user_id = u.id
            JOIN messages m ON f.message_id = m.id
            {where}
            ORDER BY f.created_at DESC, f.id DESC
            LIMIT %s OFFSET %s""",
        params + [page_size, offset],
    )
    rows = list(cursor.fetchall())

    items = []
    for row in rows:
        # 上一条学生提问：同会话中 id 小于该 assistant 消息的最近一条 user 消息
        cursor = db.execute(
            """SELECT content FROM messages
               WHERE conversation_id = %s AND id < %s AND role = 'user'
               ORDER BY id DESC LIMIT 1""",
            (row["conversation_id"], row["message_id"]),
        )
        question_row = cursor.fetchone()
        try:
            kp_ids = json.loads(row["knowledge_point_ids"]) if row["knowledge_point_ids"] else []
        except (TypeError, ValueError):
            kp_ids = []
        items.append(FeedbackItem(
            id=row["id"],
            rating=row["rating"],
            reason=row["reason"],
            student_id=row["student_id"],
            student_name=row["student_name"],
            question=question_row["content"] if question_row else None,
            answer=row["answer"],
            knowledge_point_ids=kp_ids,
            created_at=str(row["created_at"]),
        ))

    return FeedbackListOut(total=total, items=items)


# ========== 检索可观测报表（v4.0 M1） ==========

KB_STATUSES = ("ok", "empty", "timeout", "http_error", "code_error", "degraded")


@router.get("/kb/stats", response_model=KbStatsOut)
async def get_kb_stats(
    days: int = Query(7, ge=1, le=365),
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    """检索质量统计（空结果率/降级数/平均耗时/按天/按状态）"""
    time_where = "WHERE created_at >= DATE_SUB(CURDATE(), INTERVAL %s DAY)"

    cursor = db.execute(
        f"""SELECT COUNT(*) AS total,
                   SUM(status = 'empty') AS empty_count,
                   SUM(status = 'degraded') AS degraded_count,
                   AVG(elapsed_ms) AS avg_elapsed
            FROM kb_search_logs {time_where}""",
        (days,),
    )
    row = cursor.fetchone()
    # PyMySQL 聚合陷阱：SUM/AVG 返回 Decimal，需 int()/float() 包裹
    total = int(row["total"] or 0)
    empty_count = int(row["empty_count"] or 0)
    degraded_count = int(row["degraded_count"] or 0)
    avg_elapsed_ms = int(float(row["avg_elapsed"])) if row["avg_elapsed"] is not None else 0

    cursor = db.execute(
        f"""SELECT DATE(created_at) AS d, COUNT(*) AS total,
                   SUM(status = 'empty') AS empty_cnt,
                   SUM(status = 'degraded') AS degraded_cnt
            FROM kb_search_logs {time_where}
            GROUP BY DATE(created_at) ORDER BY d""",
        (days,),
    )
    by_day = [
        KbStatsByDay(
            date=str(r["d"]),  # DATE() 返回 date，需 str()
            total=int(r["total"]),
            empty=int(r["empty_cnt"] or 0),
            degraded=int(r["degraded_cnt"] or 0),
        )
        for r in cursor.fetchall()
    ]

    cursor = db.execute(
        f"""SELECT status, COUNT(*) AS cnt FROM kb_search_logs {time_where}
            GROUP BY status""",
        (days,),
    )
    by_status = {s: 0 for s in KB_STATUSES}
    for r in cursor.fetchall():
        by_status[r["status"]] = int(r["cnt"])

    return KbStatsOut(
        total=total,
        empty_count=empty_count,
        empty_rate=round(empty_count / total, 3) if total else 0.0,
        degraded_count=degraded_count,
        avg_elapsed_ms=avg_elapsed_ms,
        by_day=by_day,
        by_status=by_status,
    )


@router.get("/kb/hot-kps", response_model=list[HotKpItem])
async def get_kb_hot_kps(
    days: int = Query(30, ge=1, le=365),
    top: int = Query(10, ge=1, le=100),
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    """高频知识点 TopN（kp_ids JSON 数组 Python 侧展开聚合，量小无性能问题）"""
    cursor = db.execute(
        """SELECT kp_ids FROM kb_search_logs
           WHERE created_at >= DATE_SUB(CURDATE(), INTERVAL %s DAY)
             AND kp_ids IS NOT NULL""",
        (days,),
    )
    counter: Counter = Counter()
    for row in cursor.fetchall():
        try:
            kp_list = json.loads(row["kp_ids"])
        except (TypeError, ValueError):
            continue
        if isinstance(kp_list, list):
            counter.update(str(kp) for kp in kp_list)

    return [
        HotKpItem(kp_id=kp_id, count=count)
        for kp_id, count in counter.most_common(top)
    ]
